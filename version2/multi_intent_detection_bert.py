import torch
from torch.utils.data import Dataset, DataLoader
from transformers import BertTokenizer, BertForSequenceClassification
from sklearn.model_selection import train_test_split
from sklearn.preprocessing import MultiLabelBinarizer
from torch.optim import AdamW
from torch.nn import BCEWithLogitsLoss
from tqdm import tqdm
import pandas as pd
import os
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment

# ====================
# 数据集类
# ====================
class IntentDataset(Dataset):
    def __init__(self, texts, labels, tokenizer, max_len):
        self.texts = texts
        self.labels = labels
        self.tokenizer = tokenizer
        self.max_len = max_len

    def __len__(self):
        return len(self.texts)

    def __getitem__(self, idx):
        text = self.texts[idx]
        label = self.labels[idx]

        # Tokenize the text
        encoding = self.tokenizer(
            text,
            truncation=True,
            padding="max_length",
            max_length=self.max_len,
            return_tensors="pt",
        )
        input_ids = encoding["input_ids"].squeeze()
        attention_mask = encoding["attention_mask"].squeeze()

        return {
            "input_ids": input_ids,
            "attention_mask": attention_mask,
            "label": torch.tensor(label, dtype=torch.float),  # Multi-label: use float
        }

# ====================
# 模型相关函数
# ====================
def train_model(train_loader, val_loader, model, tokenizer, device, epochs, save_path):
    # 1 加载评估文件
    wb = load_workbook('test.xlsx')

    sheet_name = "bert-base-chinese"
    if sheet_name not in wb.sheetnames:
        ws = wb.create_sheet(title=sheet_name)
    else:
        ws = wb[sheet_name]

    # 设置表头
    headers = ["Epoch", "Train_Loss", "Train_Accuracy", "Validation_Loss", "Validation_Accuracy"]
    if ws.max_row == 1 and ws['A1'].value is None:
        ws.append(headers)

    # 设置表头样式
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center')

    optimizer = AdamW(model.parameters(), lr=2e-5)
    criterion = BCEWithLogitsLoss()  # Multi-label loss

    for epoch in range(epochs):
        model.train()
        train_loss = 0
        correct = 0
        total = 0
        for batch in tqdm(train_loader, desc=f"Training Epoch {epoch + 1}"):
            input_ids = batch["input_ids"].to(device)
            attention_mask = batch["attention_mask"].to(device)
            labels = batch["label"].to(device)

            optimizer.zero_grad()

            outputs = model(
                input_ids=input_ids,
                attention_mask=attention_mask,
                labels=None,  # No need for label in output
            )
            logits = outputs.logits
            loss = criterion(logits, labels)  # Calculate loss

            loss.backward()
            optimizer.step()

            train_loss += loss.item()

            # Binary classification per label
            preds = (torch.sigmoid(logits) > 0.5).float()
            correct += (preds == labels).sum().item()
            total += labels.size(0) * labels.size(1)  # Total count of labels

        train_loss_avg = train_loss / len(train_loader)
        train_accuracy = correct / total * 100
        print(f"Epoch {epoch + 1} Train Loss: {train_loss_avg}")
        print(f"Epoch {epoch + 1} Train Accuracy: {train_accuracy:.2f}%")

        # 验证集
        model.eval()
        val_loss = 0
        correct = 0
        total = 0
        with torch.no_grad():
            for batch in val_loader:
                input_ids = batch["input_ids"].to(device)
                attention_mask = batch["attention_mask"].to(device)
                labels = batch["label"].to(device)

                outputs = model(
                    input_ids=input_ids,
                    attention_mask=attention_mask,
                    labels=None,  # No need for label in output
                )
                logits = outputs.logits
                loss = criterion(logits, labels)
                val_loss += loss.item()

                preds = (torch.sigmoid(logits) > 0.5).float()
                correct += (preds == labels).sum().item()
                total += labels.size(0) * labels.size(1)

        val_loss_avg = val_loss / len(val_loader)
        val_accuracy = correct / total * 100
        print(f"Validation Loss: {val_loss_avg}")
        print(f"Validation Accuracy: {val_accuracy:.2f}%")

        # 保存模型和分词器
        model.save_pretrained(save_path)
        tokenizer.save_pretrained(save_path)

    print("数据文件已保存")
    print(f"模型已保存至 {save_path}")

def load_model(model_path, device):
    """加载保存的模型和分词器"""
    model = BertForSequenceClassification.from_pretrained(model_path)
    tokenizer = BertTokenizer.from_pretrained(model_path)
    label_encoder = torch.load(os.path.join(model_path, "label_mapping.pth"))
    model.to(device)
    return model, tokenizer, label_encoder

def predict(question, model, tokenizer, label_encoder, device, max_len, top_k=24):
    """对单条输入进行预测并返回概率较大的前k个标签"""
    encoding = tokenizer(
        question,
        truncation=True,
        padding="max_length",
        max_length=max_len,
        return_tensors="pt",
    )
    input_ids = encoding["input_ids"].to(device)
    attention_mask = encoding["attention_mask"].to(device)

    model.eval()
    with torch.no_grad():
        outputs = model(input_ids=input_ids, attention_mask=attention_mask)
        logits = outputs.logits
        probs = torch.sigmoid(logits).cpu().numpy()  # 获取每个标签的概率

    # 获取概率较大的前top_k个标签
    top_k_indices = probs.argsort()[0, -top_k:][::-1]  # 返回最大值索引
    top_k_probs = probs[0, top_k_indices]
    top_k_labels = label_encoder.classes_[top_k_indices]

    return [(label, prob) for label, prob in zip(top_k_labels, top_k_probs)]

# ====================
# 数据处理函数
# ====================
def prepare_data(data_path, test_size=0.2, max_len=128, model_path=None):
    """加载数据并生成数据集和 DataLoader"""
    df = pd.read_excel(data_path)
    texts = df["question"].values
    labels = df["intent"].values

    # 标签编码（使用MultiLabelBinarizer）
    mlb = MultiLabelBinarizer()
    encoded_labels = mlb.fit_transform([[label] for label in labels])  # 假设每个标签是单标签

    # 划分训练和测试集
    train_texts, val_texts, train_labels, val_labels = train_test_split(
        texts, encoded_labels, test_size=test_size, random_state=42
    )

    # 初始化分词器
    tokenizer = BertTokenizer.from_pretrained(model_path)

    # 创建数据集
    train_dataset = IntentDataset(train_texts, train_labels, tokenizer, max_len)
    val_dataset = IntentDataset(val_texts, val_labels, tokenizer, max_len)

    # 创建 DataLoader
    train_loader = DataLoader(train_dataset, batch_size=16, shuffle=True)
    val_loader = DataLoader(val_dataset, batch_size=16)

    return train_loader, val_loader, tokenizer, mlb

# ====================
# 封装的核心函数
# ====================
def start(data_path, save_path, device, model_path, max_len=128, epochs=3):
    """加载数据、训练模型并保存"""
    os.makedirs(save_path, exist_ok=True)

    # 加载数据
    train_loader, val_loader, tokenizer, mlb = prepare_data(
        data_path, max_len=max_len, model_path=model_path
    )

    # 初始化模型
    model = BertForSequenceClassification.from_pretrained(
        model_path, num_labels=len(mlb.classes_)  # 使用动态路径
    )
    model.to(device)

    # 保存标签编码器
    torch.save(mlb, os.path.join(save_path, "label_mapping.pth"))

    # 训练模型
    train_model(train_loader, val_loader, model, tokenizer, device, epochs, save_path)

def predict_intent(question, model_path, device, max_len=128):
    """加载模型并预测意图"""
    model, tokenizer, mlb = load_model(model_path, device)
    return predict(question, model, tokenizer, mlb, device, max_len)

# ====================
# 主函数入口
# ====================
if __name__ == "__main__":
    data_path = "../data/intent_detection.xlsx"  # 数据路径
    save_path = "../saved_model"  # 模型保存路径
    model_path = "D:/pycharm/example/nlp/KG/doc/model/bert-base-chinese"  # 模型路径（可以根据需要修改）
    max_len = 128
    device = torch.device("cuda" if torch.cuda.is_available() else "cpu")

    # 启动训练
    # start(data_path, save_path, device, model_path, max_len, epochs=3)

    # 测试预测
    while True:
        question = input("请输入问句（输入'退出'结束）：")
        if question == "退出":
            break
        intent_probs = predict_intent(question, save_path, device, max_len)
        print(f"预测意图及概率：{intent_probs}")
