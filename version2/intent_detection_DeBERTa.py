import pandas as pd
import torch
from torch.utils.data import DataLoader, Dataset
from transformers import DebertaTokenizer, DebertaForSequenceClassification, AdamW, DebertaV2Tokenizer, \
    DebertaV2ForSequenceClassification
from sklearn.model_selection import train_test_split
from sklearn.metrics import accuracy_score
from torch.optim.lr_scheduler import StepLR
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment

# 1. 数据加载
data_path = "data/intent_detection.xlsx"
df = pd.read_excel(data_path)

# 2. 数据预处理
tokenizer = DebertaV2Tokenizer.from_pretrained(r"D:/pycharm/example/nlp/doc/model/DeBERTa-base")


class IntentDetectionDataset(Dataset):
    def __init__(self, questions, labels, tokenizer, max_len):
        self.questions = questions
        self.labels = labels
        self.tokenizer = tokenizer
        self.max_len = max_len

    def __len__(self):
        return len(self.questions)

    def __getitem__(self, item):
        question = self.questions[item]
        label = self.labels[item]

        encoding = self.tokenizer.encode_plus(
            question,
            add_special_tokens=True,
            max_length=self.max_len,
            padding='max_length',
            truncation=True,
            return_tensors='pt'
        )

        return {
            'input_ids': encoding['input_ids'].flatten(),
            'attention_mask': encoding['attention_mask'].flatten(),
            'labels': torch.tensor(label, dtype=torch.long)
        }


# 划分训练集和验证集
X_train, X_val, y_train, y_val = train_test_split(df['question'], df['intent_id'], test_size=0.2, random_state=42)

# 创建数据集
train_dataset = IntentDetectionDataset(X_train.tolist(), y_train.tolist(), tokenizer, max_len=128)
val_dataset = IntentDetectionDataset(X_val.tolist(), y_val.tolist(), tokenizer, max_len=128)

train_dataloader = DataLoader(train_dataset, batch_size=16, shuffle=True)
val_dataloader = DataLoader(val_dataset, batch_size=16)

# 3. 加载模型
# model = DebertaForSequenceClassification.from_pretrained(r"D:/pycharm/example/nlp/doc/model/DeBERTa-base",
#                                                          num_labels=len(df['intent_id'].unique()))
model = DebertaV2ForSequenceClassification.from_pretrained(r"D:/pycharm/example/nlp/doc/model/DeBERTa-base",
                                                           num_labels=len(df['intent_id'].unique()))
device = torch.device('cuda' if torch.cuda.is_available() else 'cpu')
model.to(device)

# 4. 训练模型
optimizer = AdamW(model.parameters(), lr=2e-5)
scheduler = StepLR(optimizer, step_size=1, gamma=0.9)
criterion = torch.nn.CrossEntropyLoss()


def train_epoch(model, data_loader, optimizer, criterion):
    model.train()
    losses = []
    correct_preds = 0
    total_preds = 0

    for batch in data_loader:
        input_ids = batch['input_ids'].to(device)
        attention_mask = batch['attention_mask'].to(device)
        labels = batch['labels'].to(device)

        optimizer.zero_grad()

        outputs = model(input_ids=input_ids, attention_mask=attention_mask, labels=labels)
        loss = outputs.loss
        logits = outputs.logits

        loss.backward()
        optimizer.step()

        preds = torch.argmax(logits, dim=1)
        correct_preds += torch.sum(preds == labels)
        total_preds += len(labels)

        losses.append(loss.item())

    scheduler.step()

    avg_loss = sum(losses) / len(losses)
    accuracy = correct_preds.double() / total_preds
    return avg_loss, accuracy


def eval_model(model, data_loader):
    model.eval()
    correct_preds = 0
    total_preds = 0
    all_preds = []
    all_labels = []

    with torch.no_grad():
        for batch in data_loader:
            input_ids = batch['input_ids'].to(device)
            attention_mask = batch['attention_mask'].to(device)
            labels = batch['labels'].to(device)

            outputs = model(input_ids=input_ids, attention_mask=attention_mask)
            logits = outputs.logits

            preds = torch.argmax(logits, dim=1)
            correct_preds += torch.sum(preds == labels)
            total_preds += len(labels)

            all_preds.extend(preds.cpu().numpy())
            all_labels.extend(labels.cpu().numpy())

    accuracy = correct_preds.double() / total_preds
    return accuracy, all_preds, all_labels

# 1 加载评估文件
wb = load_workbook('test.xlsx')

sheet_name = "DeBERTa"
if sheet_name not in wb.sheetnames:
    ws = wb.create_sheet(title=sheet_name)
else:
    ws = wb[sheet_name]

# 设置表头
headers = ["Epoch", "Train_Loss", "Train_Accuracy",  "Validation_Accuracy"]
if ws.max_row == 1 and ws['A1'].value is None:
    ws.append(headers)

# 设置表头样式
for cell in ws[1]:
    cell.font = Font(bold=True)
    cell.alignment = Alignment(horizontal='center')

# 训练过程
epochs = 15
for epoch in range(epochs):
    print(f"Epoch {epoch + 1}/{epochs}")

    # 训练
    train_loss, train_accuracy = train_epoch(model, train_dataloader, optimizer, criterion)
    print(f"Train loss: {train_loss:.4f}, Train accuracy: {train_accuracy:.4f}")

    # 验证
    val_accuracy, val_preds, val_labels = eval_model(model, val_dataloader)
    print(f"Validation accuracy: {val_accuracy:.4f}")
     # 将每个epoch的数据写入Excel
    # 将 torch.Tensor 转换为 Python 的 float 类型
    ws.append([
        epoch,
        train_loss,
        train_accuracy.item(),  # 转换为 float
        val_accuracy.item()  # 转换为 float
    ])

wb.save('test.xlsx')

# 5. 保存模型
model.save_pretrained("saved_model_deberta")
tokenizer.save_pretrained("saved_model_deberta")

# 6. 使用while语句进行实时预测
model.eval()
while True:
    question = input("请输入问题（输入'退出'结束）：")
    if question == '退出':
        break

    inputs = tokenizer.encode_plus(
        question,
        add_special_tokens=True,
        max_length=128,
        padding='max_length',
        truncation=True,
        return_tensors='pt'
    )

    input_ids = inputs['input_ids'].to(device)
    attention_mask = inputs['attention_mask'].to(device)

    with torch.no_grad():
        outputs = model(input_ids=input_ids, attention_mask=attention_mask)
        logits = outputs.logits
        pred = torch.argmax(logits, dim=1).item()

    intent = df[df['intent_id'] == pred]['intent'].values[0]
    print(f"预测的意图是：{intent}")
