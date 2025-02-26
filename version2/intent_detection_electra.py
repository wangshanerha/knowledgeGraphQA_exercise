import pandas as pd
import torch
from torch.utils.data import DataLoader, Dataset
from transformers import ElectraTokenizer, ElectraForSequenceClassification, AdamW
from sklearn.model_selection import train_test_split
from sklearn.metrics import accuracy_score
from torch.utils.tensorboard import SummaryWriter
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment

# 数据路径
data_path = "data/intent_detection.xlsx"

# 加载数据
df = pd.read_excel(data_path)
data = df[["question", "intent_id"]].rename(columns={"question": "text", "intent_id": "label"})

# 数据集定义
class IntentDataset(Dataset):
    def __init__(self, texts, labels, tokenizer, max_length=128):
        self.texts = texts
        self.labels = labels
        self.tokenizer = tokenizer
        self.max_length = max_length

    def __len__(self):
        return len(self.texts)

    def __getitem__(self, idx):
        text = str(self.texts[idx])
        label = self.labels[idx]
        encoding = self.tokenizer.encode_plus(
            text,
            add_special_tokens=True,
            max_length=self.max_length,
            padding="max_length",
            truncation=True,
            return_tensors="pt"
        )
        return {
            "input_ids": encoding["input_ids"].flatten(),
            "attention_mask": encoding["attention_mask"].flatten(),
            "labels": torch.tensor(label, dtype=torch.long)
        }

# 加载tokenizer和模型
model_path = "D:/pycharm/example/nlp/doc/model/electra"  # 使用指定路径
tokenizer = ElectraTokenizer.from_pretrained(model_path)
model = ElectraForSequenceClassification.from_pretrained(model_path, num_labels=len(data['label'].unique()))

# 数据划分
train_texts, val_texts, train_labels, val_labels = train_test_split(data['text'].values, data['label'].values, test_size=0.2)

# 创建DataLoader
train_dataset = IntentDataset(train_texts, train_labels, tokenizer)
val_dataset = IntentDataset(val_texts, val_labels, tokenizer)

train_loader = DataLoader(train_dataset, batch_size=16, shuffle=True)
val_loader = DataLoader(val_dataset, batch_size=16)

# 设置优化器
optimizer = AdamW(model.parameters(), lr=2e-5)

# 设置TensorBoard日志
writer = SummaryWriter()

# 训练函数
def train_epoch(model, data_loader, optimizer, device):
    model = model.train()
    losses = []
    correct_predictions = 0

    for batch in data_loader:
        optimizer.zero_grad()

        input_ids = batch["input_ids"].to(device)
        attention_mask = batch["attention_mask"].to(device)
        labels = batch["labels"].to(device)

        outputs = model(input_ids=input_ids, attention_mask=attention_mask, labels=labels)
        loss = outputs.loss
        logits = outputs.logits

        loss.backward()
        optimizer.step()

        losses.append(loss.item())
        preds = torch.argmax(logits, dim=1)
        correct_predictions += torch.sum(preds == labels)

    accuracy = correct_predictions.double() / len(data_loader.dataset)
    return accuracy, sum(losses) / len(losses)

# 验证函数
def eval_epoch(model, data_loader, device):
    model = model.eval()
    losses = []
    correct_predictions = 0

    with torch.no_grad():
        for batch in data_loader:
            input_ids = batch["input_ids"].to(device)
            attention_mask = batch["attention_mask"].to(device)
            labels = batch["labels"].to(device)

            outputs = model(input_ids=input_ids, attention_mask=attention_mask, labels=labels)
            loss = outputs.loss
            logits = outputs.logits

            losses.append(loss.item())
            preds = torch.argmax(logits, dim=1)
            correct_predictions += torch.sum(preds == labels)

    accuracy = correct_predictions.double() / len(data_loader.dataset)
    return accuracy, sum(losses) / len(losses)

# 设置设备
device = torch.device("cuda" if torch.cuda.is_available() else "cpu")
model = model.to(device)

# 训练过程
epochs =30
best_accuracy = 0
# #  加载评估文件
# wb = load_workbook('test.xlsx')
#
# sheet_name = "electra"
# if sheet_name not in wb.sheetnames:
#     ws = wb.create_sheet(title=sheet_name)
# else:
#     ws = wb[sheet_name]
#
# # 设置表头
# headers = ["Epoch", "Train_Loss", "Train_Accuracy", "Validation_Loss", "Validation_Accuracy"]
# if ws.max_row == 1 and ws['A1'].value is None:
#     ws.append(headers)
#
# # 设置表头样式
# for cell in ws[1]:
#     cell.font = Font(bold=True)
#     cell.alignment = Alignment(horizontal='center')

for epoch in range(epochs):
    print(f"Epoch {epoch + 1}/{epochs}...")
    train_accuracy, train_loss = train_epoch(model, train_loader, optimizer, device)
    val_accuracy, val_loss = eval_epoch(model, val_loader, device)

    print(f"Train Accuracy: {train_accuracy:.4f}, Train Loss: {train_loss:.4f}")
    print(f"Val Accuracy: {val_accuracy:.4f}, Val Loss: {val_loss:.4f}")

    # 记录训练和验证数据
    writer.add_scalar('Train Accuracy', train_accuracy, epoch)
    writer.add_scalar('Train Loss', train_loss, epoch)
    writer.add_scalar('Validation Accuracy', val_accuracy, epoch)
    writer.add_scalar('Validation Loss', val_loss, epoch)

    # 保存最佳模型
    if val_accuracy > best_accuracy:
        best_accuracy = val_accuracy
        print(f"Best Validation accuracy {best_accuracy:.4f}.")
        model.save_pretrained("./saved_model_electra")
        tokenizer.save_pretrained("./saved_model_electra")
     # 将每个epoch的数据写入Excel
#     ws.append([
#         epoch,
#         train_loss,
#         train_accuracy.item(),  # 转换为 float
#         val_loss,
#         val_accuracy.item()  # 转换为 float
#     ])
#
# wb.save('test.xlsx')

# 关闭TensorBoard
writer.close()

# 模型预测函数
def predict(text, model, tokenizer, device):
    model = model.eval()
    encoding = tokenizer.encode_plus(
        text,
        add_special_tokens=True,
        max_length=128,
        padding="max_length",
        truncation=True,
        return_tensors="pt"
    )
    input_ids = encoding["input_ids"].to(device)
    attention_mask = encoding["attention_mask"].to(device)

    with torch.no_grad():
        outputs = model(input_ids=input_ids, attention_mask=attention_mask)
        logits = outputs.logits
        preds = torch.argmax(logits, dim=1).item()

    return preds

# 启动while循环进行预测
while True:
    text = input("请输入文本进行预测（输入'quit'退出）：")
    if text.lower() == "quit":
        print("退出预测程序。")
        break

    # 调用预测函数
    intent = predict(text, model, tokenizer, device)
    print(f"预测的意图ID：{intent}")