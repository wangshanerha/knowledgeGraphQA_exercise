import torch
import torch.nn as nn
import torch.optim as optim
from torch.utils.data import DataLoader, Dataset
from sklearn.model_selection import train_test_split
from sklearn.preprocessing import LabelEncoder
from sklearn.metrics import accuracy_score, precision_score, recall_score, f1_score
import pandas as pd
import numpy as np
import os
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, Alignment

# ===================== 配置参数 =====================
FILE_PATH = "../../data/intent_detection.xlsx"
SAVE_PATH = "../../saved_models/Intent_Recognition_TextCNN"
EXCEL_PATH = "../../assess/Intent_Recognition_TextCNN.xlsx"
MAX_LEN = 128
BATCH_SIZE = 16
EMBED_DIM = 50
KERNEL_SIZES = [3, 4, 5]
NUM_FILTERS = 16
EPOCHS = 20
LEARNING_RATE = 0.001

# ===================== Excel初始化 =====================
def init_excel():
    os.makedirs(os.path.dirname(EXCEL_PATH), exist_ok=True)
    
    if os.path.exists(EXCEL_PATH):
        wb = load_workbook(EXCEL_PATH)
        if "Metrics" in wb.sheetnames:
            ws = wb["Metrics"]
        else:
            ws = wb.create_sheet("Metrics")
    else:
        wb = Workbook()
        ws = wb.active
        ws.title = "Metrics"
        headers = ["Epoch", "Train_Loss", "Train_Accuracy", 
                 "Val_Loss", "Val_Accuracy", "Val_Precision", 
                 "Val_Recall", "Val_F1"]
        ws.append(headers)
        # 设置表头样式
        for col in range(1, len(headers)+1):
            cell = ws.cell(row=1, column=col)
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal='center')
    
    return wb, ws

# ===================== 数据加载 =====================
def load_dataset(file_path):
    data = pd.read_excel(file_path)
    texts = data['question'].tolist()
    labels = data['intent'].tolist()
    return texts, labels

# ===================== 数据集类 =====================
class TextDataset(Dataset):
    def __init__(self, texts, labels, tokenizer, label_encoder, max_len):
        self.texts = texts
        self.labels = label_encoder.transform(labels)
        self.tokenizer = tokenizer
        self.max_len = max_len

    def __len__(self):
        return len(self.texts)

    def __getitem__(self, idx):
        text = self.texts[idx]
        tokens = self.tokenizer(text)
        # 补齐/截断序列
        if len(tokens) < self.max_len:
            tokens += [0] * (self.max_len - len(tokens))
        else:
            tokens = tokens[:self.max_len]
        return torch.tensor(tokens), torch.tensor(self.labels[idx])

# ===================== 简单分词器 =====================
def simple_tokenizer(text):
    vocab = {char: idx+1 for idx, char in enumerate(set("".join(texts)))}
    return [vocab.get(char, 0) for char in text]

# ===================== TextCNN模型 =====================
class TextCNN(nn.Module):
    def __init__(self, vocab_size, embed_dim, num_classes, kernel_sizes, num_filters):
        super(TextCNN, self).__init__()
        self.embedding = nn.Embedding(vocab_size, embed_dim)
        self.convs = nn.ModuleList([
            nn.Conv2d(1, num_filters, (k, embed_dim)) for k in kernel_sizes
        ])
        self.fc = nn.Linear(num_filters * len(kernel_sizes), num_classes)
        self.dropout = nn.Dropout(0.5)

    def forward(self, x):
        x = self.embedding(x)  # (batch_size, seq_len, embed_dim)
        x = x.unsqueeze(1)      # (batch_size, 1, seq_len, embed_dim)
        
        # 卷积和池化
        conv_outputs = []
        for conv in self.convs:
            conv_out = torch.relu(conv(x)).squeeze(3)  # (batch_size, num_filters, seq_len-k+1)
            pool_out = torch.max_pool1d(conv_out, conv_out.size(2)).squeeze(2)  # (batch_size, num_filters)
            conv_outputs.append(pool_out)
        
        x = torch.cat(conv_outputs, 1)  # (batch_size, num_filters * len(kernel_sizes))
        x = self.dropout(x)
        return self.fc(x)

# ===================== 训练函数 =====================
def train_epoch(model, dataloader, optimizer, criterion, device):
    model.train()
    total_loss = 0.0
    correct = 0
    total = 0
    
    for inputs, targets in dataloader:
        inputs, targets = inputs.to(device), targets.to(device)
        
        optimizer.zero_grad()
        outputs = model(inputs)
        loss = criterion(outputs, targets)
        loss.backward()
        optimizer.step()
        
        total_loss += loss.item()
        _, preds = torch.max(outputs, 1)
        correct += (preds == targets).sum().item()
        total += targets.size(0)
    
    return total_loss / len(dataloader), correct / total

# ===================== 验证函数 =====================
def evaluate(model, dataloader, criterion, device):
    model.eval()
    total_loss = 0.0
    all_preds = []
    all_labels = []
    
    with torch.no_grad():
        for inputs, targets in dataloader:
            inputs, targets = inputs.to(device), targets.to(device)
            
            outputs = model(inputs)
            loss = criterion(outputs, targets)
            
            total_loss += loss.item()
            _, preds = torch.max(outputs, 1)
            
            all_preds.extend(preds.cpu().numpy())
            all_labels.extend(targets.cpu().numpy())
    
    metrics = {
        'loss': total_loss / len(dataloader),
        'accuracy': accuracy_score(all_labels, all_preds),
        'precision': precision_score(all_labels, all_preds, average='macro', zero_division=0),
        'recall': recall_score(all_labels, all_preds, average='macro', zero_division=0),
        'f1': f1_score(all_labels, all_preds, average='macro', zero_division=0)
    }
    return metrics

if __name__ == "__main__":
    # 初始化Excel
    wb, ws = init_excel()
    
    # 加载数据
    texts, labels = load_dataset(FILE_PATH)  # labels是原始字符串
    
    # 标签编码（仅fit，保持原始标签）
    label_encoder = LabelEncoder()
    label_encoder.fit(labels)  # 重要：在数据分割前fit
    
    # 数据分割（使用原始标签进行分层）
    train_texts, val_texts, train_labels, val_labels = train_test_split(
        texts, 
        labels,  # 原始字符串标签
        test_size=0.2, 
        random_state=42, 
        stratify=labels  # 正确分层抽样
    )
    
    # 创建数据集（传入原始标签）
    vocab_size = len(set("".join(texts))) + 1
    train_dataset = TextDataset(train_texts, train_labels, simple_tokenizer, label_encoder, MAX_LEN)
    val_dataset = TextDataset(val_texts, val_labels, simple_tokenizer, label_encoder, MAX_LEN)
    
    # 创建DataLoader
    train_loader = DataLoader(train_dataset, batch_size=BATCH_SIZE, shuffle=True)
    val_loader = DataLoader(val_dataset, batch_size=BATCH_SIZE)
    
    # 设备设置
    device = torch.device("cuda" if torch.cuda.is_available() else "cpu")
    
    # 模型初始化
    model = TextCNN(
        vocab_size=vocab_size,
        embed_dim=EMBED_DIM,
        num_classes=len(label_encoder.classes_),
        kernel_sizes=KERNEL_SIZES,
        num_filters=NUM_FILTERS
    ).to(device)
    
    # 损失函数与优化器
    criterion = nn.CrossEntropyLoss()
    optimizer = optim.Adam(model.parameters(), lr=LEARNING_RATE)
    
    # 创建保存目录
    os.makedirs(SAVE_PATH, exist_ok=True)
    
    # 训练循环
    best_f1 = 0.0
    for epoch in range(EPOCHS):
        # 训练
        train_loss, train_acc = train_epoch(model, train_loader, optimizer, criterion, device)
        
        # 验证
        val_metrics = evaluate(model, val_loader, criterion, device)
        
        # 记录到Excel
        ws.append([
            epoch + 1,
            train_loss,
            train_acc,
            val_metrics['loss'],
            val_metrics['accuracy'],
            val_metrics['precision'],
            val_metrics['recall'],
            val_metrics['f1']
        ])
        wb.save(EXCEL_PATH)
        
        # 保存最佳模型
        if val_metrics['f1'] > best_f1:
            best_f1 = val_metrics['f1']
            torch.save({
                'model_state_dict': model.state_dict(),
                'label_encoder': label_encoder
            }, os.path.join(SAVE_PATH, 'best_model.pth'))
        
        # 打印日志
        print(f"\nEpoch {epoch+1}/{EPOCHS}")
        print(f"Train Loss: {train_loss:.4f} | Acc: {train_acc:.4f}")
        print(f"Val Loss: {val_metrics['loss']:.4f} | Acc: {val_metrics['accuracy']:.4f}")
        print(f"Precision: {val_metrics['precision']:.4f} | Recall: {val_metrics['recall']:.4f} | F1: {val_metrics['f1']:.4f}")

    print(f"\nTraining completed. Best model saved to: {os.path.join(SAVE_PATH, 'best_model.pth')}")
    print(f"Metrics saved to: {EXCEL_PATH}")