import pandas as pd
import torch
import os
from torch.utils.data import Dataset, DataLoader
from transformers import AutoTokenizer, AutoModelForSequenceClassification, AdamW
from sklearn.model_selection import train_test_split
from sklearn.preprocessing import LabelEncoder
from sklearn.metrics import accuracy_score, precision_score, recall_score, f1_score
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, Alignment

# ====================
# 配置参数
# ====================
MODEL_NAME = "../../models/DeBERTa-base"
DATA_PATH = "../../data/intent_detection.xlsx"
SAVE_PATH = "../../saved_models/Intent_Recognition_DeBERTa"
EXCEL_PATH = "../../assess/Intent_Recognition_DeBERTa.xlsx"
MAX_LENGTH = 128
BATCH_SIZE = 16
EPOCHS = 10
DEVICE = torch.device("cuda" if torch.cuda.is_available() else "cpu")

# ====================
# 初始化Excel记录
# ====================
def init_excel():
    os.makedirs(os.path.dirname(EXCEL_PATH), exist_ok=True)
    
    if os.path.exists(EXCEL_PATH):
        wb = load_workbook(EXCEL_PATH)
        if "Metrics" in wb.sheetnames:
            ws = wb["Metrics"]
        else:
            ws = wb.create_sheet("Metrics")
    else:
        wb = Workbook()
        ws = wb.active
        ws.title = "Metrics"
        headers = ["Epoch", "Train_Loss", "Train_Accuracy", 
                   "Val_Loss", "Val_Accuracy", "Val_Precision", 
                   "Val_Recall", "Val_F1"]
        ws.append(headers)
        # 设置表头样式
        for col in range(1, len(headers)+1):
            cell = ws.cell(row=1, column=col)
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal='center')
    
    return wb, ws

# ====================
# 自定义Dataset
# ====================
class IntentDataset(Dataset):
    def __init__(self, texts, labels, tokenizer, max_length):
        self.texts = texts
        self.labels = labels
        self.tokenizer = tokenizer
        self.max_length = max_length

    def __len__(self):
        return len(self.texts)

    def __getitem__(self, idx):
        encoding = self.tokenizer(
            str(self.texts[idx]),
            max_length=self.max_length,
            padding='max_length',
            truncation=True,
            return_tensors='pt'
        )
        return {
            'input_ids': encoding['input_ids'].squeeze(),
            'attention_mask': encoding['attention_mask'].squeeze(),
            'labels': torch.tensor(self.labels[idx], dtype=torch.long)
        }

# ====================
# 主流程
# ====================
def main():
    # 初始化Excel
    wb, ws = init_excel()
    
    # 加载数据
    df = pd.read_excel(DATA_PATH)
    df = df[['question', 'intent']].dropna()
    
    # 标签编码
    le = LabelEncoder()
    df['label'] = le.fit_transform(df['intent'])
    
    # 数据集划分（添加分层抽样）
    train_df, val_df = train_test_split(
        df,
        test_size=0.2,
        stratify=df['intent'],  # 分层抽样
        random_state=42
    )
    
    # 初始化模型和分词器
    tokenizer = AutoTokenizer.from_pretrained(MODEL_NAME)
    model = AutoModelForSequenceClassification.from_pretrained(
        MODEL_NAME,
        num_labels=len(le.classes_),
        ignore_mismatched_sizes=True  # 忽略尺寸不匹配警告
    ).to(DEVICE)
    
    # 创建DataLoader
    train_dataset = IntentDataset(train_df['question'].values, train_df['label'].values, tokenizer, MAX_LENGTH)
    val_dataset = IntentDataset(val_df['question'].values, val_df['label'].values, tokenizer, MAX_LENGTH)
    train_loader = DataLoader(train_dataset, batch_size=BATCH_SIZE, shuffle=True)
    val_loader = DataLoader(val_dataset, batch_size=BATCH_SIZE)
    
    # 优化器
    optimizer = AdamW(model.parameters(), lr=2e-5)
    
    # 训练参数
    best_accuracy = 0.0
    
    for epoch in range(EPOCHS):
        # 训练阶段
        model.train()
        total_loss, correct, total = 0.0, 0, 0
        for batch in train_loader:
            optimizer.zero_grad()
            inputs = {k: v.to(DEVICE) for k, v in batch.items()}
            outputs = model(**inputs)
            
            loss = outputs.loss
            loss.backward()
            optimizer.step()
            
            total_loss += loss.item()
            preds = torch.argmax(outputs.logits, dim=1)
            correct += (preds == inputs["labels"]).sum().item()
            total += inputs["labels"].size(0)
        
        # 计算训练指标
        train_loss = total_loss / len(train_loader)
        train_acc = correct / total
        
        # 验证阶段
        model.eval()
        val_loss, all_preds, all_labels = 0.0, [], []
        with torch.no_grad():
            for batch in val_loader:
                inputs = {k: v.to(DEVICE) for k, v in batch.items()}
                outputs = model(**inputs)
                
                val_loss += outputs.loss.item()
                preds = torch.argmax(outputs.logits, dim=1)
                all_preds.extend(preds.cpu().numpy())
                all_labels.extend(inputs["labels"].cpu().numpy())
        
        # 计算验证指标
        val_loss_avg = val_loss / len(val_loader)
        val_acc = accuracy_score(all_labels, all_preds)
        val_precision = precision_score(all_labels, all_preds, average='macro', zero_division=0)
        val_recall = recall_score(all_labels, all_preds, average='macro', zero_division=0)
        val_f1 = f1_score(all_labels, all_preds, average='macro', zero_division=0)
        
        # 记录到Excel
        ws.append([
            epoch + 1,
            train_loss,
            train_acc,
            val_loss_avg,
            val_acc,
            val_precision,
            val_recall,
            val_f1
        ])
        wb.save(EXCEL_PATH)
        
        # 保存最佳模型
        if val_acc > best_accuracy:
            best_accuracy = val_acc
            model.save_pretrained(SAVE_PATH)
            tokenizer.save_pretrained(SAVE_PATH)
            print(f"New best model saved to {SAVE_PATH}")
        
        # 打印日志
        print(f"\nEpoch {epoch+1}/{EPOCHS}")
        print(f"Train Loss: {train_loss:.4f} | Acc: {train_acc:.4f}")
        print(f"Val Loss: {val_loss_avg:.4f} | Acc: {val_acc:.4f}")
        print(f"Precision: {val_precision:.4f} | Recall: {val_recall:.4f} | F1: {val_f1:.4f}")

if __name__ == "__main__":
    # 创建保存目录
    os.makedirs(SAVE_PATH, exist_ok=True)
    main()
