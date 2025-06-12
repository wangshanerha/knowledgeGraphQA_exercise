import pandas as pd
import torch
import os
from torch.utils.data import DataLoader, Dataset
from transformers import ElectraTokenizer, ElectraForSequenceClassification, AdamW
from sklearn.model_selection import train_test_split
from sklearn.metrics import accuracy_score, precision_score, recall_score, f1_score
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, Alignment

# ====================
# 配置参数
# ====================
DATA_PATH = "../../data/intent_detection.xlsx"
MODEL_SAVE_PATH = "../../saved_models/Intent_Recognition_ELECTRA"
EXCEL_PATH = "../../assess/Intent_Recognition_ELECTRA.xlsx"
INIT_MODEL_PATH = "../../models/chinese-electra-large-generator"

# ====================
# 数据集类
# ====================
class IntentDataset(Dataset):
    def __init__(self, texts, labels, tokenizer, max_length=128):
        self.texts = texts
        self.labels = labels
        self.tokenizer = tokenizer
        self.max_length = max_length

    def __len__(self):
        return len(self.texts)

    def __getitem__(self, idx):
        encoding = self.tokenizer.encode_plus(
            str(self.texts[idx]),
            add_special_tokens=True,
            max_length=self.max_length,
            padding="max_length",
            truncation=True,
            return_tensors="pt"
        )
        return {
            "input_ids": encoding["input_ids"].squeeze(),
            "attention_mask": encoding["attention_mask"].squeeze(),
            "labels": torch.tensor(self.labels[idx], dtype=torch.long)
        }

# ====================
# 初始化Excel记录
# ====================
def init_excel():
    os.makedirs(os.path.dirname(EXCEL_PATH), exist_ok=True)
    
    if os.path.exists(EXCEL_PATH):
        wb = load_workbook(EXCEL_PATH)
        if "Metrics" in wb.sheetnames:
            ws = wb["Metrics"]
        else:
            ws = wb.create_sheet("Metrics")
    else:
        wb = Workbook()
        ws = wb.active
        ws.title = "Metrics"
        headers = ["Epoch", "Train_Loss", "Train_Accuracy", 
                   "Val_Loss", "Val_Accuracy", "Val_Precision", 
                   "Val_Recall", "Val_F1"]
        ws.append(headers)
        
        # 设置表头样式
        for col in range(1, len(headers)+1):
            cell = ws.cell(row=1, column=col)
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal='center')
    
    return wb, ws

# ====================
# 验证函数（含完整指标）
# ====================
def eval_epoch(model, data_loader, device):
    model.eval()
    losses = []
    all_preds = []
    all_labels = []

    with torch.no_grad():
        for batch in data_loader:
            inputs = {k: v.to(device) for k, v in batch.items()}
            outputs = model(**inputs)
            
            loss = outputs.loss.item()
            preds = torch.argmax(outputs.logits, dim=1)
            
            losses.append(loss)
            all_preds.extend(preds.cpu().numpy())
            all_labels.extend(inputs["labels"].cpu().numpy())

    avg_loss = sum(losses) / len(losses)
    accuracy = accuracy_score(all_labels, all_preds)
    precision = precision_score(all_labels, all_preds, average='macro', zero_division=0)
    recall = recall_score(all_labels, all_preds, average='macro', zero_division=0)
    f1 = f1_score(all_labels, all_preds, average='macro', zero_division=0)
    
    return avg_loss, accuracy, precision, recall, f1

# ====================
# 主流程
# ====================
def main():
    # 初始化Excel
    wb, ws = init_excel()
    
    # 加载数据
    df = pd.read_excel(DATA_PATH)
    data = df[["question", "intent_id"]].rename(columns={"question": "text", "intent_id": "label"})
    
    # 初始化模型
    tokenizer = ElectraTokenizer.from_pretrained(INIT_MODEL_PATH)
    model = ElectraForSequenceClassification.from_pretrained(
        INIT_MODEL_PATH, 
        num_labels=len(data['label'].unique())
    )
    
    # 划分数据集
    train_texts, val_texts, train_labels, val_labels = train_test_split(
        data['text'].values, 
        data['label'].values, 
        test_size=0.2,
        stratify=data['label'].values  # 添加分层抽样
    )
    
    # 创建DataLoader
    train_dataset = IntentDataset(train_texts, train_labels, tokenizer)
    val_dataset = IntentDataset(val_texts, val_labels, tokenizer)
    train_loader = DataLoader(train_dataset, batch_size=16, shuffle=True)
    val_loader = DataLoader(val_dataset, batch_size=16)
    
    # 设备设置
    device = torch.device("cuda" if torch.cuda.is_available() else "cpu")
    model = model.to(device)
    optimizer = AdamW(model.parameters(), lr=2e-5)
    
    # 训练参数
    epochs = 10
    best_accuracy = 0.0
    
    for epoch in range(epochs):
        # 训练阶段
        model.train()
        train_loss, correct, total = 0.0, 0, 0
        
        for batch in train_loader:
            optimizer.zero_grad()
            inputs = {k: v.to(device) for k, v in batch.items()}
            outputs = model(**inputs)
            
            loss = outputs.loss
            loss.backward()
            optimizer.step()
            
            train_loss += loss.item()
            preds = torch.argmax(outputs.logits, dim=1)
            correct += (preds == inputs["labels"]).sum().item()
            total += inputs["labels"].size(0)
        
        # 计算训练指标
        train_loss_avg = train_loss / len(train_loader)
        train_acc = correct / total
        
        # 验证阶段
        val_loss, val_acc, val_precision, val_recall, val_f1 = eval_epoch(model, val_loader, device)
        
        # 记录到Excel
        ws.append([
            epoch + 1,
            train_loss_avg,
            train_acc,
            val_loss,
            val_acc,
            val_precision,
            val_recall,
            val_f1
        ])
        wb.save(EXCEL_PATH)
        
        # 保存最佳模型
        if val_acc > best_accuracy:
            best_accuracy = val_acc
            model.save_pretrained(MODEL_SAVE_PATH)
            tokenizer.save_pretrained(MODEL_SAVE_PATH)
            print(f"New best model saved to {MODEL_SAVE_PATH}")
        
        # 打印日志
        print(f"\nEpoch {epoch+1}/{epochs}")
        print(f"Train Loss: {train_loss_avg:.4f} | Acc: {train_acc:.4f}")
        print(f"Val Loss: {val_loss:.4f} | Acc: {val_acc:.4f}")
        print(f"Precision: {val_precision:.4f} | Recall: {val_recall:.4f} | F1: {val_f1:.4f}")

if __name__ == "__main__":
    # 创建保存目录
    os.makedirs(MODEL_SAVE_PATH, exist_ok=True)
    main()