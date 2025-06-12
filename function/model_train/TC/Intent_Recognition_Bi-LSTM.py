import torch
import torch.nn as nn
import pandas as pd
import os
import numpy as np
from torch.utils.data import Dataset, DataLoader
from torch.nn.utils.rnn import pack_padded_sequence, pad_packed_sequence
from sklearn.model_selection import train_test_split
from sklearn.preprocessing import LabelEncoder
from sklearn.metrics import accuracy_score, precision_score, recall_score, f1_score
from transformers import BertTokenizer
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, Alignment

# ===================== 配置参数 =====================
FILEPATH = '../../data/intent_detection.xlsx'
SAVE_PATH = '../../saved_models/Intent_Recognition_BiLSTM'
EXCEL_PATH = '../../assess/Intent_Recognition_BiLSTM.xlsx'
MAX_LENGTH = 128
BATCH_SIZE = 32
EMBEDDING_DIM = 256
HIDDEN_DIM = 256
N_LAYERS = 2
DROPOUT = 0.3
EPOCHS = 10
LEARNING_RATE = 0.001

# ===================== Excel初始化 =====================
def init_excel():
    os.makedirs(os.path.dirname(EXCEL_PATH), exist_ok=True)
    
    if os.path.exists(EXCEL_PATH):
        wb = load_workbook(EXCEL_PATH)
        if "Metrics" in wb.sheetnames:
            ws = wb["Metrics"]
        else:
            ws = wb.create_sheet("Metrics")
    else:
        wb = Workbook()
        ws = wb.active
        ws.title = "Metrics"
        headers = ["Epoch", "Train_Loss", "Train_Accuracy", 
                   "Val_Loss", "Val_Accuracy", "Val_Precision", 
                   "Val_Recall", "Val_F1"]
        ws.append(headers)
        for col in range(1, len(headers)+1):
            cell = ws.cell(row=1, column=col)
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal='center')
    
    return wb, ws

# ===================== 数据预处理 =====================
def load_and_preprocess_data(filepath, max_length):
    df = pd.read_excel(filepath)
    df = df[['question', 'intent']].dropna()
    
    # 标签编码
    le = LabelEncoder()
    labels = le.fit_transform(df['intent'])
    
    # 分词处理
    tokenizer = BertTokenizer.from_pretrained('../../models/bert-base-chinese')
    
    # 处理文本并记录长度
    tokenized_texts = []
    lengths = []
    for text in df['question']:
        encoded = tokenizer.encode(text, add_special_tokens=True, max_length=max_length, truncation=True)
        tokenized_texts.append(encoded)
        lengths.append(len(encoded))
    
    # 补齐序列
    padded_texts = np.array([seq + [0]*(max_length - len(seq)) for seq in tokenized_texts])
    return padded_texts, np.array(lengths), labels, le

# ===================== 数据集类 =====================
class IntentDataset(Dataset):
    def __init__(self, texts, lengths, labels):
        self.texts = texts
        self.lengths = lengths
        self.labels = labels
        
    def __len__(self):
        return len(self.texts)
    
    def __getitem__(self, idx):
        return {
            'text': torch.tensor(self.texts[idx], dtype=torch.long),
            'length': self.lengths[idx],
            'label': torch.tensor(self.labels[idx], dtype=torch.long)
        }

# ===================== 原始BiLSTM模型 =====================
class OriginalBiLSTM(nn.Module):
    def __init__(self, vocab_size, embedding_dim, hidden_dim, num_classes):
        super(OriginalBiLSTM, self).__init__()
        
        # 嵌入层（保持与改进模型相同的初始化）
        self.embedding = nn.Embedding(vocab_size, embedding_dim)
        
        # BiLSTM层（参数完全一致）
        self.lstm = nn.LSTM(embedding_dim, hidden_dim, 
                          num_layers=N_LAYERS,
                          bidirectional=True,
                          batch_first=True,
                          dropout=DROPOUT if N_LAYERS > 1 else 0)
        
        # 分类层（保持相同结构）
        self.fc = nn.Sequential(
            nn.Dropout(DROPOUT),
            nn.Linear(hidden_dim*2, num_classes)  # 双向隐藏层拼接
        )
    
    def forward(self, text, lengths):
        # 嵌入层
        embedded = self.embedding(text)
        
        # 打包序列（保留变长序列处理）
        packed = pack_padded_sequence(embedded, lengths, 
                                    batch_first=True, 
                                    enforce_sorted=False)
        
        # LSTM处理
        packed_out, (hidden, cell) = self.lstm(packed)
        
        # 双向LSTM的隐藏状态拼接：前向和后向的最后时刻状态
        hidden = torch.cat((hidden[-2], hidden[-1]), dim=1)
        
        # 分类输出
        return self.fc(hidden)

# ===================== 训练函数 =====================
def train_epoch(model, dataloader, optimizer, criterion, device):
    model.train()
    total_loss = 0.0
    correct = 0
    total = 0
    
    for batch in dataloader:
        texts = batch['text'].to(device)
        labels = batch['label'].to(device)
        lengths = batch['length'].to('cpu')
        
        optimizer.zero_grad()
        
        outputs = model(texts, lengths)
        loss = criterion(outputs, labels)
        
        loss.backward()
        optimizer.step()
        
        total_loss += loss.item()
        _, preds = torch.max(outputs, 1)
        correct += (preds == labels).sum().item()
        total += labels.size(0)
    
    return total_loss / len(dataloader), correct / total

# ===================== 验证函数 =====================
def evaluate(model, dataloader, criterion, device):
    model.eval()
    total_loss = 0.0
    all_preds = []
    all_labels = []
    
    with torch.no_grad():
        for batch in dataloader:
            texts = batch['text'].to(device)
            labels = batch['label'].to(device)
            lengths = batch['length'].to('cpu')
            
            outputs = model(texts, lengths)
            loss = criterion(outputs, labels)
            
            total_loss += loss.item()
            _, preds = torch.max(outputs, 1)
            
            all_preds.extend(preds.cpu().numpy())
            all_labels.extend(labels.cpu().numpy())
    
    metrics = {
        'loss': total_loss / len(dataloader),
        'accuracy': accuracy_score(all_labels, all_preds),
        'precision': precision_score(all_labels, all_preds, average='macro', zero_division=0),
        'recall': recall_score(all_labels, all_preds, average='macro', zero_division=0),
        'f1': f1_score(all_labels, all_preds, average='macro', zero_division=0)
    }
    return metrics

# ===================== 主流程 =====================
if __name__ == "__main__":
    # 初始化Excel
    wb, ws = init_excel()
    
    # 数据准备
    texts, lengths, labels, label_encoder = load_and_preprocess_data(FILEPATH, MAX_LENGTH)
    texts_train, texts_val, lengths_train, lengths_val, labels_train, labels_val = train_test_split(
        texts, lengths, labels,
        test_size=0.2,
        stratify=labels,
        random_state=42
    )
    
    # 创建DataLoader
    train_dataset = IntentDataset(texts_train, lengths_train, labels_train)
    val_dataset = IntentDataset(texts_val, lengths_val, labels_val)
    
    train_loader = DataLoader(train_dataset, batch_size=BATCH_SIZE, shuffle=True)
    val_loader = DataLoader(val_dataset, batch_size=BATCH_SIZE)
    
    # 设备设置
    device = torch.device('cuda' if torch.cuda.is_available() else 'cpu')
    
    # 模型初始化
    tokenizer = BertTokenizer.from_pretrained('../../models/bert-base-chinese')
    model = OriginalBiLSTM(
        vocab_size=tokenizer.vocab_size,
        embedding_dim=EMBEDDING_DIM,
        hidden_dim=HIDDEN_DIM,
        num_classes=len(label_encoder.classes_)
    ).to(device)
    
    # 优化器和损失函数
    optimizer = torch.optim.Adam(model.parameters(), lr=LEARNING_RATE)
    criterion = nn.CrossEntropyLoss()
    
    # 创建保存目录
    os.makedirs(SAVE_PATH, exist_ok=True)
    
    # 训练循环
    best_f1 = 0.0
    for epoch in range(EPOCHS):
        # 训练
        train_loss, train_acc = train_epoch(model, train_loader, optimizer, criterion, device)
        
        # 验证
        val_metrics = evaluate(model, val_loader, criterion, device)
        
        # 记录到Excel
        ws.append([
            epoch + 1,
            train_loss,
            train_acc,
            val_metrics['loss'],
            val_metrics['accuracy'],
            val_metrics['precision'],
            val_metrics['recall'],
            val_metrics['f1']
        ])
        wb.save(EXCEL_PATH)
        
        # 保存最佳模型
        if val_metrics['f1'] > best_f1:
            best_f1 = val_metrics['f1']
            torch.save({
                'model_state_dict': model.state_dict(),
                'label_encoder': label_encoder
            }, os.path.join(SAVE_PATH, 'best_model.pth'))
        
        # 打印日志
        print(f"\nEpoch {epoch+1}/{EPOCHS}")
        print(f"Train Loss: {train_loss:.4f} | Acc: {train_acc:.4f}")
        print(f"Val Loss: {val_metrics['loss']:.4f} | Acc: {val_metrics['accuracy']:.4f}")
        print(f"Precision: {val_metrics['precision']:.4f} | Recall: {val_metrics['recall']:.4f} | F1: {val_metrics['f1']:.4f}")
    
    print(f"\nTraining completed. Best model saved to: {os.path.join(SAVE_PATH, 'best_model.pth')}")
    print(f"Training metrics saved to: {EXCEL_PATH}")