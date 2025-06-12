import torch
from torch.utils.data import Dataset, DataLoader
from transformers import BertTokenizer, BertForSequenceClassification
from sklearn.model_selection import train_test_split
from sklearn.preprocessing import LabelEncoder
from sklearn.metrics import precision_score, recall_score, f1_score
from torch.optim import AdamW
from tqdm import tqdm
import pandas as pd
import numpy as np
import os
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, Alignment

# ====================
# 数据集类
# ====================
class IntentDataset(Dataset):
    def __init__(self, texts, labels, tokenizer, max_len):
        self.texts = texts
        self.labels = labels
        self.tokenizer = tokenizer
        self.max_len = max_len

    def __len__(self):
        return len(self.texts)

    def __getitem__(self, idx):
        encoding = self.tokenizer(
            self.texts[idx],
            truncation=True,
            padding="max_length",
            max_length=self.max_len,
            return_tensors="pt",
        )
        return {
            "input_ids": encoding["input_ids"].squeeze(),
            "attention_mask": encoding["attention_mask"].squeeze(),
            "label": torch.tensor(self.labels[idx], dtype=torch.long),
        }

# ====================
# 模型训练函数（修复参数传递问题）
# ====================
def train_model(train_loader, val_loader, model, tokenizer, device, epochs, save_path):
    # 初始化Excel文件
    excel_path = "../../assess/Intent_Recognition_BERT.xlsx"
    os.makedirs(os.path.dirname(excel_path), exist_ok=True)
    
    if os.path.exists(excel_path):
        wb = load_workbook(excel_path)
        if "Training Metrics" in wb.sheetnames:
            ws = wb["Training Metrics"]
        else:
            ws = wb.create_sheet("Training Metrics")
    else:
        wb = Workbook()
        ws = wb.active
        ws.title = "Training Metrics"
        headers = ["Epoch", "Train_Loss", "Train_Accuracy", 
                   "Val_Loss", "Val_Accuracy", "Val_Precision", 
                   "Val_Recall", "Val_F1"]
        ws.append(headers)
        # 设置表头样式
        for col in range(1, len(headers)+1):
            cell = ws.cell(row=1, column=col)
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal='center')

    optimizer = AdamW(model.parameters(), lr=2e-5)
    best_accuracy = 0.0

    for epoch in range(epochs):
        # 训练阶段
        model.train()
        train_loss, correct, total = 0.0, 0, 0
        for batch in tqdm(train_loader, desc=f"Training Epoch {epoch + 1}"):
            # 显式获取各参数
            input_ids = batch["input_ids"].to(device)
            attention_mask = batch["attention_mask"].to(device)
            labels = batch["label"].to(device)

            # 前向传播
            outputs = model(
                input_ids=input_ids,
                attention_mask=attention_mask,
                labels=labels  # 使用正确的参数名称
            )
            
            loss = outputs.loss
            loss.backward()
            optimizer.step()
            optimizer.zero_grad()

            # 计算指标
            train_loss += loss.item()
            preds = torch.argmax(outputs.logits, dim=1)
            correct += (preds == labels).sum().item()
            total += labels.size(0)

        # 计算训练指标
        train_loss_avg = train_loss / len(train_loader)
        train_acc = correct / total

        # 验证阶段
        model.eval()
        val_loss, all_preds, all_labels = 0.0, [], []
        with torch.no_grad():
            for batch in val_loader:
                input_ids = batch["input_ids"].to(device)
                attention_mask = batch["attention_mask"].to(device)
                labels = batch["label"].to(device)

                outputs = model(
                    input_ids=input_ids,
                    attention_mask=attention_mask,
                    labels=labels
                )
                
                val_loss += outputs.loss.item()
                preds = torch.argmax(outputs.logits, dim=1)
                all_preds.extend(preds.cpu().numpy())
                all_labels.extend(labels.cpu().numpy())

        # 计算验证指标
        val_loss_avg = val_loss / len(val_loader)
        val_acc = np.mean(np.array(all_preds) == np.array(all_labels))
        val_precision = precision_score(all_labels, all_preds, average='macro', zero_division=0)
        val_recall = recall_score(all_labels, all_preds, average='macro', zero_division=0)
        val_f1 = f1_score(all_labels, all_preds, average='macro', zero_division=0)

        # 写入Excel
        ws.append([
            epoch + 1,
            train_loss_avg,
            train_acc,
            val_loss_avg,
            val_acc,
            val_precision,
            val_recall,
            val_f1
        ])
        wb.save(excel_path)

        # 保存最佳模型
        if val_acc > best_accuracy:
            best_accuracy = val_acc
            model.save_pretrained(save_path)
            tokenizer.save_pretrained(save_path)
            print(f"New best model saved with accuracy {val_acc:.4f}")

        # 打印日志
        print(f"\nEpoch {epoch+1}/{epochs}")
        print(f"Train Loss: {train_loss_avg:.4f} | Acc: {train_acc:.4f}")
        print(f"Val Loss: {val_loss_avg:.4f} | Acc: {val_acc:.4f}")
        print(f"Precision: {val_precision:.4f} | Recall: {val_recall:.4f} | F1: {val_f1:.4f}")

    print(f"\nTraining metrics saved to {excel_path}")
    print(f"Final model saved to {save_path}")

# ====================
# 数据准备函数
# ====================
def prepare_data(data_path, model_path, max_len=128):
    """加载数据并生成数据集"""
    df = pd.read_excel(data_path)
    texts = df["question"].values
    labels = df["intent"].values

    # 标签编码
    label_encoder = LabelEncoder()
    encoded_labels = label_encoder.fit_transform(labels)

    # 划分数据集（添加分层抽样）
    train_texts, val_texts, train_labels, val_labels = train_test_split(
        texts,
        encoded_labels,
        test_size=0.2,
        stratify=labels,  # 分层抽样
        random_state=42
    )

    # 初始化分词器
    tokenizer = BertTokenizer.from_pretrained(model_path)

    # 创建数据集
    train_dataset = IntentDataset(train_texts, train_labels, tokenizer, max_len)
    val_dataset = IntentDataset(val_texts, val_labels, tokenizer, max_len)

    return train_dataset, val_dataset, tokenizer, label_encoder

# ====================
# 启动函数
# ====================
def start(data_path, save_path, model_path, device, max_len=128, epochs=20):
    """主启动函数"""
    os.makedirs(save_path, exist_ok=True)
    
    # 加载数据
    train_dataset, val_dataset, tokenizer, label_encoder = prepare_data(
        data_path=data_path,
        model_path=model_path,
        max_len=max_len
    )
    
    # 创建DataLoader
    train_loader = DataLoader(train_dataset, batch_size=16, shuffle=True)
    val_loader = DataLoader(val_dataset, batch_size=16)

    # 初始化模型
    model = BertForSequenceClassification.from_pretrained(
        model_path,
        num_labels=len(label_encoder.classes_)
    ).to(device)

    # 保存标签映射
    torch.save(label_encoder, os.path.join(save_path, "label_mapping.pth"))

    # 开始训练
    train_model(train_loader, val_loader, model, tokenizer, device, epochs, save_path)

# ====================
# 主函数入口
# ====================
if __name__ == "__main__":
    # 路径配置
    data_path = "../../data/intent_detection.xlsx"
    save_path = "../../saved_models/Intent_Recognition_BERT"
    model_path = "bert-base-chinese"
    device = torch.device("cuda" if torch.cuda.is_available() else "cpu")
    
    # 启动训练
    start(
        data_path=data_path,
        save_path=save_path,
        model_path=model_path,
        device=device,
        max_len=128,
        epochs=10
    )