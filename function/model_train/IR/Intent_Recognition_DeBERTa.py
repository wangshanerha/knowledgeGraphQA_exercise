import pandas as pd
import torch
from torch.utils.data import Dataset, DataLoader
from transformers import AutoTokenizer, AutoModelForSequenceClassification, AdamW
from sklearn.model_selection import train_test_split
from sklearn.preprocessing import LabelEncoder
from sklearn.metrics import accuracy_score
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment

# 参数设置
model_name = "../models/DeBERTa-base"
data_path = "../data/intent_detection.xlsx"
max_length = 128
batch_size = 16
epochs = 20
device = torch.device("cuda" if torch.cuda.is_available() else "cpu")

# 数据准备
df = pd.read_excel(data_path)
df = df[['question', 'intent']].dropna()

# 标签编码sde
le = LabelEncoder()
df['label'] = le.fit_transform(df['intent'])

# 数据集划分
train_df, val_df = train_test_split(df, test_size=0.2, random_state=42)


# 自定义Dataset
class IntentDataset(Dataset):
    def __init__(self, texts, labels, tokenizer, max_length):
        self.texts = texts
        self.labels = labels
        self.tokenizer = tokenizer
        self.max_length = max_length

    def __len__(self):
        return len(self.texts)

    def __getitem__(self, idx):
        text = str(self.texts[idx])
        encoding = self.tokenizer(
            text,
            max_length=self.max_length,
            padding='max_length',
            truncation=True,
            return_tensors='pt'
        )
        return {
            'input_ids': encoding['input_ids'].flatten(),
            'attention_mask': encoding['attention_mask'].flatten(),
            'labels': torch.tensor(self.labels[idx], dtype=torch.long)
        }


# 初始化模型和分词器
tokenizer = AutoTokenizer.from_pretrained(model_name)
model = AutoModelForSequenceClassification.from_pretrained(
    model_name,
    num_labels=len(le.classes_)
).to(device)

# 创建DataLoader
train_dataset = IntentDataset(train_df['question'].values, train_df['label'].values, tokenizer, max_length)
val_dataset = IntentDataset(val_df['question'].values, val_df['label'].values, tokenizer, max_length)

train_loader = DataLoader(train_dataset, batch_size=batch_size, shuffle=True)
val_loader = DataLoader(val_dataset, batch_size=batch_size)

# 优化器
optimizer = AdamW(model.parameters(), lr=2e-5)

# 训练循环
  # 加载评估文件
# wb = load_workbook('test.xlsx')
#
# sheet_name = "DeBERTa"
# if sheet_name not in wb.sheetnames:
#     ws = wb.create_sheet(title=sheet_name)
# else:
#     ws = wb[sheet_name]
#
# # 设置表头
# headers = ["Epoch", "Train_Loss", "avg_val_loss",  "Validation_Accuracy"]
# if ws.max_row == 1 and ws['A1'].value is None:
#     ws.append(headers)
#
# # 设置表头样式
# for cell in ws[1]:
#     cell.font = Font(bold=True)
#     cell.alignment = Alignment(horizontal='center')

for epoch in range(epochs):
    print(f"Epoch {epoch + 1}/{epochs}")
    # 训练阶段
    model.train()
    total_loss = 0
    for batch in train_loader:
        optimizer.zero_grad()
        input_ids = batch['input_ids'].to(device)
        attention_mask = batch['attention_mask'].to(device)
        labels = batch['labels'].to(device)

        outputs = model(input_ids, attention_mask=attention_mask, labels=labels)
        loss = outputs.loss
        loss.backward()
        optimizer.step()
        total_loss += loss.item()

    # 验证阶段
    model.eval()
    val_loss = 0
    predictions = []
    true_labels = []
    with torch.no_grad():
        for batch in val_loader:
            input_ids = batch['input_ids'].to(device)
            attention_mask = batch['attention_mask'].to(device)
            labels = batch['labels'].to(device)

            outputs = model(input_ids, attention_mask=attention_mask, labels=labels)
            val_loss += outputs.loss.item()

            logits = outputs.logits
            preds = torch.argmax(logits, dim=1).cpu().numpy()
            predictions.extend(preds)
            true_labels.extend(labels.cpu().numpy())

    # 输出训练信息
    avg_train_loss = total_loss / len(train_loader)
    avg_val_loss = val_loss / len(val_loader)
    val_acc = accuracy_score(true_labels, predictions)

    print(f"Epoch {epoch + 1}/{epochs}")
    print(f"Train Loss: {avg_train_loss:.4f} | Val Loss: {avg_val_loss:.4f} | Val Acc: {val_acc:.4f}")
    # # 将每个epoch的数据写入Excel
    # ws.append([
    #     epoch,
    #     avg_train_loss,
    #     avg_val_loss,
    #     val_acc
    # ])

# wb.save('test.xlsx')
# 保存模型
model.save_pretrained("saved_model_DeBERTa")
tokenizer.save_pretrained("saved_model_DeBERTa")


# 预测函数
def predict(text):
    model.eval()
    encoding = tokenizer(
        text,
        max_length=max_length,
        padding='max_length',
        truncation=True,
        return_tensors='pt'
    )
    input_ids = encoding['input_ids'].to(device)
    attention_mask = encoding['attention_mask'].to(device)

    with torch.no_grad():
        outputs = model(input_ids, attention_mask=attention_mask)

    logits = outputs.logits
    pred_label = torch.argmax(logits, dim=1).item()
    return le.inverse_transform([pred_label])[0]


while True:
        user_input = input("请输入问题（输入'exit'退出）: ")
        print((user_input))
        print(predict(user_input))
        if user_input.lower() == 'exit':
            break
        result = predict(user_input)
        print(f"预测意图: {result}\n")
